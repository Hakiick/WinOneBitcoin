import pandas as pd
from openpyxl import load_workbook
from array import array
from pyautogui import *
import keyboard
import pyautogui
from datetime import datetime
import time
import pandas as pd
from openpyxl import load_workbook
from pynput.mouse import Button, Controller

mouse = Controller()
tabTime = []
tabFinal = []






def number50():
    count = 0
    while count < 22:
        number50 = pyautogui.locateCenterOnScreen('number50.png', region=(0, 0, 1920, 1080), grayscale=False, confidence=0.97)

        if number50 is not None:
            print("number50")
            now = datetime.now()
            timestamp = datetime.timestamp(now)
            tabTime.append(timestamp)
            count = count + 1
            time.sleep(0.5)
            print(count)
            
            if count == 20:
                writer = pd.ExcelWriter('demo.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
                writer.book = load_workbook('demo.xlsx')
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                reader = pd.read_excel(r'demo.xlsx')
                tF = pd.DataFrame(data=tabTime)
                tF.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
                writer.close()
                print(tabTime)


        


#while not keyboard.is_pressed('esc'):
    number50()








